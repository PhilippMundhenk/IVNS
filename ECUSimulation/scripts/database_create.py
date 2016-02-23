#!/usr/bin/python
# -*- coding: utf-8 -*-
''' Create Table 

Added Elements:

'CyaSSL', 'Crypto_Lib_HW', 'Crypto_Lib_SW'

#===============================================================================
#     CryptoLib: AES 
#===============================================================================
..\CryptLib\With_HW_Acceleration\AES_CTR_resultTab128.xlsx
..\CryptLib\With_HW_Acceleration\AES_CTR_resultTab192.xlsx
..\CryptLib\With_HW_Acceleration\AES_CTR_resultTab256.xlsx

..\CryptLib\With_HW_Acceleration\AES_CBC_resultTab128.xlsx
..\CryptLib\With_HW_Acceleration\AES_CBC_resultTab192.xlsx
..\CryptLib\With_HW_Acceleration\AES_CBC_resultTab256.xlsx

..\CryptLib\With_HW_Acceleration\AES_ECB_resultTab128.xlsx
..\CryptLib\With_HW_Acceleration\AES_ECB_resultTab192.xlsx
..\CryptLib\With_HW_Acceleration\AES_ECB_resultTab256.xlsx

..\CryptLib\With_HW_Acceleration\AES_CMAC_resultTab128.xlsx
..\CryptLib\With_HW_Acceleration\AES_CMAC_resultTab192.xlsx
..\CryptLib\With_HW_Acceleration\AES_CMAC_resultTab256.xlsx

..\CryptLib\With_SW_Only\AES\AES_CTR_resultTab128.xlsx'
..\CryptLib\With_SW_Only\AES\AES_CTR_resultTab192.xlsx'
..\CryptLib\With_SW_Only\AES\AES_CTR_resultTab256.xlsx'

..\CryptLib\With_SW_Only\AES\AES_CBC_resultTab128.xlsx'
..\CryptLib\With_SW_Only\AES\AES_CBC_resultTab192.xlsx'
..\CryptLib\With_SW_Only\AES\AES_CBC_resultTab256.xlsx'

..\CryptLib\With_SW_Only\AES\AES_ECB_resultTab128.xlsx'
..\CryptLib\With_SW_Only\AES\AES_ECB_resultTab192.xlsx'
..\CryptLib\With_SW_Only\AES\AES_ECB_resultTab256.xlsx'

..\CryptLib\With_SW_Only\AES\AES_CMAC_resultTab128.xlsx'
..\CryptLib\With_SW_Only\AES\AES_CMAC_resultTab192.xlsx'
..\CryptLib\With_SW_Only\AES\AES_CMAC_resultTab256.xlsx'


KEYGENERATION
..\CryptLib\With_SW_Only\AES\AES_Keygeneration128 192 256.xlsx'

#===============================================================================
#     CryptoLib RSA 

Note: --- RSA E3 measure it anew! Rest is ok
#===============================================================================

..\CryptLib\With_SW_Only\RSA\Sign_Verify\E5\RSA_E5_resultTab512.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E5\RSA_E5_resultTab1024.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E5\RSA_E5_resultTab2048.xlsx

..\CryptLib\With_SW_Only\RSA\Sign_Verify\E17\RSA_E17_resultTab512.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E17\RSA_E17_resultTab1024.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E17\RSA_E17_resultTab2048.xlsx

..\CryptLib\With_SW_Only\RSA\Sign_Verify\E257\RSA_E257_resultTab512.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E257\RSA_E257_resultTab1024.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E257\RSA_E257_resultTab2048.xlsx

..\CryptLib\With_SW_Only\RSA\Sign_Verify\E65537\RSA_E65537_resultTab512.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E65537\RSA_E65537_resultTab1024.xlsx
..\CryptLib\With_SW_Only\RSA\Sign_Verify\E65537\RSA_E65537_resultTab2048.xlsx


#===============================================================================
#     Crypto Lib HASH
#===============================================================================

--- SHA1 and MD5 support HW --- not implemented yet
..\CryptLib\With_SW_Only\HASH\HASH_MD5_resultTabSW.xlsx
..\CryptLib\With_SW_Only\HASH\HASH_SHA1_resultTabSW.xlsx
..\CryptLib\With_SW_Only\HASH\HASH_SHA256_resultTabSW.xlsx


#===============================================================================
#     CryptoLib ECC
#===============================================================================

..\CryptLib\With_SW_Only\ECC\Keygen\ECC_P192_Keygen_resultTab.xlsx'
..\CryptLib\With_SW_Only\ECC\Keygen\ECC_P256_Keygen_resultTab.xlsx'
..\CryptLib\With_SW_Only\ECC\Keygen\ECC_P384_Keygen_resultTab.xlsx'

..\CryptLib\With_SW_Only\ECC\Sign_Verify\ECC_P_resultTab192.xlsx'
..\CryptLib\With_SW_Only\ECC\Sign_Verify\ECC_P_resultTab256.xlsx'
..\CryptLib\With_SW_Only\ECC\Sign_Verify\ECC_P_resultTab384.xlsx'

#===============================================================================
#     CyaSSL: AES
Note:     AES CTR Mode looks strange possibly faulty implemented?
#===============================================================================

..\CyaSSL\AES\AES_CBC_resultTab128.xlsx
..\CyaSSL\AES\AES_CBC_resultTab192.xlsx
..\CyaSSL\AES\AES_CBC_resultTab256.xlsx

..\CyaSSL\AES\AES_CCM_resultTab128.xlsx
..\CyaSSL\AES\AES_CCM_resultTab192.xlsx
..\CyaSSL\AES\AES_CCM_resultTab256.xlsx

KEYGENERATION
..\CyaSSL\AES\AES_Keygeneration128 192 256.xlsx'

#===============================================================================
#     CyaSSL: RSA - PublicEncrypt and PrivateDecrypt & Keygen // inverse operations of sign and verify!
#===============================================================================

..\CyaSSL\RSA\E3\RSA_E3_resultTab512.xlsx
..\CyaSSL\RSA\E5\RSA_E5_resultTab512.xlsx
..\CyaSSL\RSA\E17\RSA_E17_resultTab512.xlsx
..\CyaSSL\RSA\E257\RSA_E257_resultTab512.xlsx
..\CyaSSL\RSA\E65537\RSA_E65537_resultTab512.xlsx
..\CyaSSL\RSA\E3\RSA_E3_resultTab1024.xlsx
..\CyaSSL\RSA\E5\RSA_E5_resultTab1024.xlsx
..\CyaSSL\RSA\E17\RSA_E17_resultTab1024.xlsx
..\CyaSSL\RSA\E257\RSA_E257_resultTab1024.xlsx
..\CyaSSL\RSA\E65537\RSA_E65537_resultTab1024.xlsx
..\CyaSSL\RSA\E65537\RSA_E65537_resultTab2048.xlsx

..\CyaSSL\RSA\E3\RSA_E3_Keygen_resultTab512.xlsx
..\CyaSSL\RSA\E3\RSA_E3_Keygen_resultTab1024.xlsx
..\CyaSSL\RSA\E5\RSA_E5_Keygen_resultTab512.xlsx
..\CyaSSL\RSA\E5\RSA_E5_Keygen_resultTab1024.xlsx
..\CyaSSL\RSA\E17\RSA_E17_Keygen_resultTab512.xlsx
..\CyaSSL\RSA\E17\RSA_E17_Keygen_resultTab1024.xlsx
..\CyaSSL\RSA\E257\RSA_E257_Keygen_resultTab512.xlsx
..\CyaSSL\RSA\E257\RSA_E257_Keygen_resultTab1024.xlsx
..\CyaSSL\RSA\E65537\RSA_E65537_Keygen_resultTab512.xlsx
..\CyaSSL\RSA\E65537\RSA_E65537_Keygen_resultTab1024.xlsx


#===============================================================================
#     CyaSSL: ECC - Encrypt/Decrypt and Sign and Verify(HASH) & Keygen
#===============================================================================

..\CyaSSL\ECC\ECC_P256_Cryption_resultTab.xlsx    NOTE: ONLY 16, 32, 48,... measured!
..\CyaSSL\ECC\ECC_P384_Cryption_resultTab.xlsx
..\CyaSSL\ECC\ECC_P521_Cryption_resultTab.xlsx

..\CyaSSL\ECC\ECC_P256_SignVerifyHash_resultTab.xlsx
..\CyaSSL\ECC\ECC_P384_SignVerifyHash_resultTab.xlsx
..\CyaSSL\ECC\ECC_P521_SignVerifyHash_resultTab.xlsx

..\CyaSSL\ECC\ECC_P256_Keygen_resultTab.xlsx
..\CyaSSL\ECC\ECC_P384_Keygen_resultTab.xlsx
..\CyaSSL\ECC\ECC_P521_Keygen_resultTab.xlsx

#===============================================================================
#     CyaSSL HASH
#===============================================================================
..\CyaSSL\HASH\MD5_HASH_resultTab.xlsx
..\CyaSSL\HASH\SHA256_HASH_resultTab.xlsx
..\CyaSSL\HASH\SHA1_HASH_resultTab.xlsx

'''

import sqlite3 as lite
from openpyxl.reader.excel import load_workbook
con = None
con = lite.connect(r"C:\Users\artur.mrowca\workspace\ECUSimulation\config\data\measurements.db")
cur = con.cursor()
 
# cur.execute("CREATE TABLE Measurements(Library TEXT, Mode TEXT, Algorithm TEXT, AlgorithmMode TEXT, Keylength INT, Exponent INT, Parameterlength INT, Datasize INT, Time DOUBLE)")
library = "'Crypto_Lib_SW'"
algorithm = "'AES'"
alg_mode = "''"
variant = "'KEYGEN'"
Keylength = 256
exponent = 0
parameter = 0
rel_col = 5  # 5 enc 6 dec
enc_time = 0.0000275625

file_path = r'C:\Users\artur.mrowca\Desktop\Measurements\measurements\CyaSSL\HASH\SHA1_HASH_resultTab.xlsx'
 
# wb = load_workbook(file_path)
# ws = wb.active
# 
# for i in range(2, int(99999999)):
#     cel_enc = ws.cell(row=i, column=rel_col)
#     cel_byte = ws.cell(row=i, column=3)
#        
#     bytess = cel_byte.value
#     enc_time = cel_enc.value
#        
#     if bytess == None:
#         break
       
# exc_str = "INSERT OR REPLACE INTO Measurements VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)" % (library, variant, algorithm, alg_mode, Keylength, exponent, parameter, 0, enc_time)
# try:
#     cur.execute(exc_str)
# except:
#     print("ERROR AT %s" % 0)
 
# cur.execute('SELECT * FROM Measurements WHERE Mode = "KEYGEN" AND Algorithm="AES"')
# data = cur.fetchall()
# print ("SQLite: %s" % [data])        
# con.commit()
# 
# con.close()
    


