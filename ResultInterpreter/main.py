
from openpyxl.reader.excel import load_workbook
import sys

print("read in File")
file_path = r"C:\Users\artur.mrowca\Desktop\abc.xlsx"

wb = load_workbook(file_path)
ws = wb.active
print("done reading in File")
r = -1
already_ok = []
print("Hallo")
while True:
    
    r += 1
    # MonitorTags.CP_ECU_RECEIVE_GRANT_MESSAGE
    tag = ws.cell(row=r, column=3).value
    s_id = ws.cell(row=r, column=6).value
    
    if tag == None: 
        break
    
    if tag == "MonitorTags.CP_ECU_RECEIVE_GRANT_MESSAGE" and s_id not in already_ok:
        already_ok.append(s_id)

print("Number granted: %s" % len(already_ok))    
sys.exit()

import xlsxwriter
import shutil


fname = r"C:\Users\artur.mrowca\Desktop\results.txt"
# output_xlsx = r"C:\Users\artur.mrowca\Desktop\Measurements\resultTab.xlsx"


#===============================================================================

output_folder = r"\CryptLib\With_SW_Only\\"
# output_folder = r"\CryptLib\With_HW_Acceleration\\"
output_folder = r"\CyaSSL\\"

alg = "ABC"
mode = "CTR"
bit = "HW"
prescaler = 15 + 1 

#===============================================================================


deff = r"C:\Users\artur.mrowca\Desktop\Measurements\measurements"
file_name = alg + "_" + mode + "_resultTab" + bit + ".xlsx"
output_xlsx = deff + output_folder + alg + "\\" + file_name

a = deff + output_folder + alg + "\\" + file_name[:-5] + "_pre" + str(prescaler - 1) + ".txt"
shutil.copyfile(fname, a)


#===============================================================================
#  1. read from  file
#===============================================================================
with open(fname, encoding='utf-8', errors='ignore') as f:
    content = f.readlines()
    
lst_of_lsts = []
cur_dict = {}
found = False


t_one_cycle = prescaler * (1 / 16000000)

for el in content:
    
    if found:
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['algorithm'] = sub[0]
        try:
            cur_dict['key_length'] = int(sub[2])
            cur_dict['mode'] = sub[1]
        except:
            try:
                cur_dict['key_length'] = int(sub[1])
            except:
                cur_dict['key_length'] = 0
            cur_dict['mode'] = sub[0]
        
        
        found = False
    
    if el.find('test ') != -1:
        cur_dict = {}
        cur_dict['data_size'] = int(el[5:]) 
        found = True
    
    if el.find('avg_enc ') != -1:
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['avg_enc'] = int(sub[1]) 
        cur_dict['avg_enc_secs'] = int(sub[1]) * t_one_cycle
         
    if el.find('min_enc ') != -1:
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['min_enc'] = int(sub[1]) 
        cur_dict['min_enc_secs'] = int(sub[1]) * t_one_cycle
         
    if el.find('max_enc ') != -1:
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['max_enc'] = int(sub[1])  
        cur_dict['max_enc_secs'] = int(sub[1]) * t_one_cycle
         
    if el.find('avg_dec ') != -1:
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['avg_dec'] = int(sub[1]) 
        cur_dict['avg_dec_secs'] = int(sub[1]) * t_one_cycle
         
    if el.find('min_dec ') != -1:
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['min_dec'] = int(sub[1]) 
        cur_dict['min_dec_secs'] = int(sub[1]) * t_one_cycle
    
    if el.find('max_dec ') != -1: 
        sub = el.split(' ')
        sub = [i.replace("\n", '') for i in sub]
        cur_dict['max_dec'] = int(sub[1]) 
        cur_dict['max_dec_secs'] = int(sub[1]) * t_one_cycle
        lst_of_lsts.append(cur_dict)
        
        
#===============================================================================
#     Write to excel
#===============================================================================

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook(output_xlsx)
worksheet = workbook.add_worksheet()


# Start from the first cell. Rows and columns are zero indexed.
row = 2
col = 2

worksheet.write(0, 0, 'Algorithm')
worksheet.write(0, 1, 'Mode')
worksheet.write(0, 2, 'Key length')
worksheet.write(0, 3, 'Data crypted')

worksheet.write(0, 5, 'Avg Enc [sec]')
worksheet.write(0, 6, 'Avg Dec [sec]')
worksheet.write(0, 8, 'Min Enc [sec]')
worksheet.write(0, 9, 'Max Enc [sec]')
worksheet.write(0, 10, 'Min Dec [sec]')
worksheet.write(0, 11, 'Max Dec [sec]')

worksheet.write(0, 13, 'Avg Enc [cycles]')
worksheet.write(0, 14, 'Avg Dec [cycles]')
worksheet.write(0, 15, 'Min Enc [cycles]')
worksheet.write(0, 16, 'Max Enc [cycles]')
worksheet.write(0, 17, 'Min Dec [cycles]')
worksheet.write(0, 18, 'Max Dec [cycles]')


# Iterate over the data and write it out row by row.
for dic in lst_of_lsts:
    
    worksheet.write(row, 0, dic['algorithm'])
    worksheet.write(row, 1, dic['mode'])
    worksheet.write(row, 2, dic['key_length'])
    worksheet.write(row, 3, dic['data_size'])
   
    worksheet.write(row, 5, dic['avg_enc_secs'])
    worksheet.write(row, 6, dic['avg_dec_secs'])

    worksheet.write(row, 8, dic['min_enc_secs'])
    worksheet.write(row, 9, dic['max_enc_secs'])
    
    worksheet.write(row, 10, dic['min_dec_secs'])
    worksheet.write(row, 11, dic['max_dec_secs'])
    
    worksheet.write(row, 13, dic['avg_enc'])
    worksheet.write(row, 14, dic['avg_dec'])
    
    worksheet.write(row, 15, dic['min_enc'])
    worksheet.write(row, 16, dic['max_enc'])
    
    worksheet.write(row, 17, dic['min_dec'])
    worksheet.write(row, 18, dic['max_dec'])
    
    row += 1

workbook.close()


